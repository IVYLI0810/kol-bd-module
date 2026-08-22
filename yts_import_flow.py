#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YTS 流程导入：万能模板 + 表头识别 + 字段推导 + 频道ID反查

- 模板列覆盖全生命周期：运营只填到当前进度，空着=还没到
- 解析时按表头子串匹配，团队现有进度表（频道名称/频道链接/负责人/归属月份…）
  也能直接上传，识别到几列映射几列
- 频道ID：云端反查 YouTube（公司网连不上时自动兜底为 HDL_ 前缀ID，预览里标注）
"""
import io
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import requests

NOW_YEAR = datetime.now().year

# key, 模板表头, 必填, 匹配子串（小写）
COLS = [
    ("channel_url",  "频道链接*",        True,  ["频道链接", "channel_url", "youtube"]),
    ("channel_name", "频道名称*",        True,  ["频道名称", "昵称", "channel_name"]),
    ("recruiter",    "负责人*",          True,  ["负责人", "挖掘人", "recruiter"]),
    ("plan_month",   "归属月份",         False, ["归属月份", "计划上线", "plan_month"]),
    ("price",        "报价（韩币）",      False, ["报价", "price"]),
    ("category",     "核心类目",         False, ["核心类目", "类目", "垂类", "category"]),
    ("email",        "联系邮箱",         False, ["邮箱", "email"]),
    ("emailed",      "已发邮件(Y/N)",    False, ["已发邮件"]),
    ("negotiating",  "洽谈中(Y/N)",      False, ["洽谈中", "洽谈"]),
    ("guideline",    "Guideline已发(Y/N)", False, ["guideline", "guide"]),
    ("contract",     "合同已签(Y/N)",    False, ["合同"]),
    ("gmc",          "GMC校验(Y/N)",     False, ["gmc"]),
    ("ordered",      "已下单(Y/N)",      False, ["下单"]),
    ("received",     "已收货(Y/N)",      False, ["收货"]),
    ("shoot_status", "拍摄状态",         False, ["拍摄"]),
    ("video_link",   "上传视频链接",     False, ["视频链接", "video_link"]),
    ("submit_deadline", "视频上传时间",  False, ["视频上传时间", "交稿截止", "上传时间"]),
    ("audit",        "审核结果",         False, ["审核结果", "审核状态"]),
    ("recheck",      "复审链接",         False, ["复审"]),
    ("video_views",  "播放量",           False, ["播放"]),
    ("video_likes",  "点赞数",           False, ["点赞"]),
    ("video_comments", "评论数",         False, ["评论"]),
    ("product_views", "点击量",          False, ["点击"]),
    ("orders",       "成交量",           False, ["成交量", "成交"]),
    ("gmv",          "GMV",              False, ["gmv"]),
    ("closed",       "已闭环(Y/N)",      False, ["闭环"]),
    ("notes",        "备注",             False, ["备注", "notes"]),
]

_EXAMPLES = [
    # 刚开始（挖掘/洽谈阶段，不填归属月份/报价）
    ["https://youtube.com/@example_beauty", "예시채널A", "艾薇李", "", "",
     "뷰티", "a@example.com", "Y", "Y", "", "", "", "", "", "", "", "", "",
     "", "", "", "", "", "", "", "", "新洽谈，待确认合作"],
    # 履约中（三分支完成、已下单、已填报价）
    ["https://youtube.com/@example_home", "예시채널B", "崔士杰", "2026-09",
     1500000, "홈&가든", "", "Y", "", "Y", "Y", "Y", "Y", "Y", "拍摄中", "",
     "2026-09-21", "", "", "", "", "", "", "", "", "", ""],
    # 已闭环（回填数据+报价）
    ["https://youtube.com/@example_pet", "예시채널C", "梁泳妍", "2026-08",
     2000000, "애완동물", "", "Y", "Y", "Y", "Y", "Y", "Y", "Y", "已完成",
     "https://youtube.com/shorts/xxx", "2026-08-21", "已通过", "",
     120000, 5400, 210, 8000, 96, 2100, "Y", "闭环示例"],
]


def build_template_bytes(roster=None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook()
    ws = wb.active
    ws.title = "流程导入"
    ws.append([c[1] for c in COLS])
    for row in _EXAMPLES:
        ws.append(row)
    if roster:
        formula = '"' + ",".join(roster) + '"'
        if len(formula) <= 250:  # Excel 下拉列表公式上限
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.errorTitle = "负责人不在名单"
            dv.error = "请从下拉里选名字；新成员先在挖掘站登记，再重新下载模板"
            ws.add_data_validation(dv)
            dv.add("C2:C500")
    ws2 = wb.create_sheet("填写说明")
    for line in [
        "带*为必填；Y/N 列填 Y 或留空；归属月份形如 2026-09 或 9月。",
        "只填到当前进度即可：刚洽谈就只填前几列，已闭环就把数据列补满。",
        "报价（韩币）填纯数字（如 1500000），导入后直接显示在履约详情的报价卡。",
        "两种视频链接别填混：「上传视频链接」= 正式发布链接，已闭环的填正式上线的"
        "YouTube 链接（闭环以它为准）；还在审核中的先填当前审核链接，审核通过后"
        "在网页闭环节点改填正式链接即可。驳回后重新审核的链接另填「复审链接」。",
        "「已闭环(Y/N)」填 Y：该网红直接进 📊 分析模块追踪数据，"
        "不再出现在履约中；记得把播放/点赞/评论/点击/成交/GMV 补满。",
        "负责人列带下拉（名单与挖掘站实时同步）；新成员请先到挖掘站登记，"
        "再重新下载模板。手写名字也可以，导入时会自动模糊匹配到名单。",
        "团队现有进度表也可直接上传，能识别的列（频道名称/频道链接/负责人/"
        "归属月份/视频上传时间/核心类目等）会自动映射。",
    ]:
        ws2.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
def yn(v) -> bool:
    return str(v or "").strip().lower() in ("y", "yes", "是", "1", "true", "✅")


def norm_month(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    m = re.match(r"^(20\d{2})[-年/.](\d{1,2})月?$", s)
    if m:
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            return f"{m.group(1)}-{mo:02d}"
        return s  # 月份超界：原样返回，由调用方校验拦截并提示
    m = re.match(r"^(\d{1,2})月$", s)
    if m:
        mo = int(m.group(1))
        if 1 <= mo <= 12:
            return f"{NOW_YEAR}-{mo:02d}"
        return s
    return s


def norm_date(v) -> str:
    """Excel 日期格/文本 → YYYY-MM-DD"""
    if v is None or str(v).strip() == "":
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(20\d{2})[-年/.](\d{1,2})[-月/.](\d{1,2})日?$", s)
    if m:
        mo, d = int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{m.group(1)}-{mo:02d}-{d:02d}"
        return s
    # 不带年份的运营常用写法：9月21日 → 当年
    m = re.match(r"^(\d{1,2})月(\d{1,2})日?$", s)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{NOW_YEAR}-{mo:02d}-{d:02d}"
        return s
    return s


def _num(v):
    try:
        f = float(str(v).replace(",", "").strip())
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return None


def _match_header(cell) -> str:
    s = str(cell or "").strip().lower().replace("*", "")
    if not s:
        return ""
    for key, _hdr, _req, subs in COLS:
        for sub in subs:
            if sub.lower() in s:
                return key
    return ""


def parse_workbook(data: bytes):
    """返回 (rows, issues)：rows=[{key: 原始值}], issues=[str]"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    grid = [[c.value for c in row] for row in ws.iter_rows(max_row=15)]
    head_idx, col_map = -1, {}
    for i, row in enumerate(grid):
        mapped = {}
        for j, cell in enumerate(row):
            k = _match_header(cell)
            if k and k not in mapped:
                mapped[k] = j
        if "channel_url" in mapped and "channel_name" in mapped:
            head_idx, col_map = i, mapped
            break
    if head_idx < 0:
        return [], ["未识别到表头：请至少包含「频道链接」「频道名称」两列，"\
                    "或直接使用下载的模板"]
    rows, issues = [], []
    # 遍历表头之后的所有行
    for row in list(ws.iter_rows(min_row=head_idx + 2)):
        vals = [c.value for c in row]
        rec = {}
        for key, j in col_map.items():
            if j < len(vals) and vals[j] not in (None, ""):
                rec[key] = vals[j]
        url = str(rec.get("channel_url") or "").strip().lower()
        name = str(rec.get("channel_name") or "").strip()
        if url and not url.startswith(("http", "www.", "@", "youtube",
                                       "m.youtube")):
            rec.pop("channel_url", None)
            url = ""
        if not url and not name:
            continue
        if not url:
            issues.append(f"行「{name}」缺频道链接，已跳过")
            continue
        if not rec.get("recruiter"):
            issues.append(f"行「{name}」缺负责人，已跳过")
            continue
        rows.append(rec)
    return rows, issues


# ---------------------------------------------------------------------------
def _fetch_channel_id(url: str):
    try:
        r = requests.get(url, timeout=8, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"})
        t = r.text
    except Exception:
        return None
    m = re.search(r'"channelId"\s*:\s*"(UC[\w-]{20,})"', t) \
        or re.search(r"channelId=(UC[\w-]{20,})", t) \
        or re.search(r'itemprop="identifier"[^>]*content="(UC[\w-]{20,})"', t)
    return m.group(1) if m else None


def _fallback_id(url: str) -> str:
    m = re.search(r"youtube\.com/@([\w.-]+)", url) \
        or re.search(r"youtube\.com/(?:channel/|c/)?([\w.-]+)", url)
    slug = m.group(1) if m else re.sub(r"\W+", "_", url)[-24:]
    return "HDL_" + slug


def _norm_refs(*parts):
    """把频道ID/主页链接/handle 归一成可比较的标识集合（去重用）。"""
    refs = set()
    for s in parts:
        if not s:
            continue
        s = str(s).strip().lower()
        if not s:
            continue
        s = re.sub(r"^https?://", "", s)
        s = re.sub(r"^www\.", "", s)
        s = s.rstrip("/")
        s = re.sub(r"^(youtube\.com|youtu\.be)/", "", s)
        m = re.match(r"^(?:channel/|c/|user/)?(@?[a-z0-9._-]+)$", s)
        if m:
            refs.add(m.group(1).lstrip("@"))
    return refs


def resolve_ids(rows: list, existing_by_url: dict):
    """url -> (channel_id, resolved:bool, is_existing:bool)。

    防幽灵重复记录：同一频道在库里可能以另一种链接写法存在
    （@handle vs channel/UCxxx），按归一化标识先匹配已有记录；
    反查得到的 UC ID 若已在库里（链接写法不同）也复用；
    都匹配不上才用 HDL_ 别名兜底。"""
    out = {}
    todo = []
    existing_refs = [(cid, _norm_refs(cid, url))
                     for url, cid in existing_by_url.items()]
    existing_cids = set(existing_by_url.values())
    for r in rows:
        url = str(r["channel_url"]).strip()
        if url in existing_by_url:
            out[url] = (existing_by_url[url], True, True)
            continue
        # 链接写法不同但指向同一频道：归一化匹配已有记录
        urefs = _norm_refs(url)
        hit = next((cid for cid, refs in existing_refs if urefs & refs), "")
        if hit:
            out[url] = (hit, True, True)
            continue
        todo.append(url)
    with ThreadPoolExecutor(max_workers=6) as ex:
        for url, cid in zip(todo, ex.map(_fetch_channel_id, todo)):
            if cid:
                # 反查到的 UC ID 已在库里（URL 写法不同）→ 复用，不新建
                out[url] = (cid, True, cid in existing_cids)
            else:
                out[url] = (_fallback_id(url), False, False)
    return out


# ---------------------------------------------------------------------------
def derive_record(raw: dict, channel_id: str) -> dict:
    rec = {
        "channel_id": channel_id,
        "channel_url": str(raw.get("channel_url", "")).strip(),
        "channel_name": str(raw.get("channel_name", "")).strip(),
        "recruiter": str(raw.get("recruiter", "")).strip(),
    }
    plan = norm_month(raw.get("plan_month"))
    if plan:
        rec["plan_month"] = plan
    pr = _num(raw.get("price"))
    if pr:
        rec["price"] = int(pr)
    for key in ("category", "email", "video_link", "recheck"):
        if raw.get(key):
            rec[key if key != "recheck" else "recheck_video_url"] = \
                str(raw[key]).strip()
    if raw.get("email"):
        rec["email"] = str(raw["email"]).strip()
    if yn(raw.get("emailed")):
        rec["email_status"] = "已发送"
    if yn(raw.get("negotiating")):
        rec["stage"] = "洽谈中"
    if yn(raw.get("guideline")):
        rec["guideline_status"] = "已发送"
    if yn(raw.get("contract")):
        rec["contract_status"] = "已签"
    if yn(raw.get("gmc")):
        rec["gmc_status"] = "校验通过"
    if yn(raw.get("ordered")):
        rec["order_status"] = "已下单"
    if yn(raw.get("received")):
        rec["order_status"] = "已收货"
    if raw.get("shoot_status"):
        rec["shoot_status"] = str(raw["shoot_status"]).strip()
    dl = norm_date(raw.get("submit_deadline"))
    if dl:
        rec["submit_deadline"] = dl
    audit = str(raw.get("audit") or "").strip()
    if audit in ("已通过", "未通过", "待审核", "复审中"):
        rec["audit_status"] = audit
    elif audit in ("驳回", "不通过"):
        rec["audit_status"] = "未通过"
    # 注意：有视频链接≠送审。只有「审核结果」列明确填了「待审核」才进审核站，
    # 否则只存链接，由运营在履约详情页手动「提交审核」
    for key in ("video_views", "video_likes", "video_comments",
                "product_views", "orders", "gmv"):
        n = _num(raw.get(key))
        if n is not None:
            rec[key] = n
    if yn(raw.get("closed")):
        rec["stage"] = "已完成"
    elif plan:
        rec["stage"] = "已确认"
    if raw.get("notes"):
        rec["notes"] = str(raw["notes"]).strip()
    return rec
